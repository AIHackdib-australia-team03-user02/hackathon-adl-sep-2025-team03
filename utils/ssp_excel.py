# utils/ssp_excel.py
from __future__ import annotations
import openpyxl, os, shutil, time
from datetime import datetime

# --- Force columns (your template) ---
# D = 4 (Identifier), I = 9 (P applicability)
FORCE_IDENTIFIER_COL = 4
FORCE_ISP_COL = 9

# Header matching (still used to find P/Q/R columns by name)
HEADER_HINTS = {
    "identifier":    ["identifier", "control id", "id"],
    "description":   ["description", "control description"],
    "resp_entity":   ["responsible entity", "owner", "responsible", "entity"],
    "impl_status":   ["implementation status", "status"],
    "impl_comments": ["implementation comments", "comments", "rationale", "remediation"],
    "is_p":          ["p applicability", "protected applicability", "applicability (p)", "p relevant", "protected relevant", "relevant to p"],
}

def _norm(s): 
    return str(s or "").strip().lower()

def _safe_copy_for_read(path):
    """Copy workbook to a temp file to avoid OneDrive/Excel locks while reading."""
    base, ext = os.path.splitext(path)
    tmp = f"{base}.__tmpread__{int(time.time())}{ext}"
    try:
        shutil.copy2(path, tmp)
    except PermissionError:
        time.sleep(0.7)
        shutil.copy2(path, tmp)
    return tmp

def _timestamped_path(out_path):
    base, ext = os.path.splitext(out_path)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{base}_{ts}{ext}"

def safe_save(wb, out_path):
    """Always save to a new timestamped file to avoid write locks."""
    target = _timestamped_path(out_path)
    wb.save(target)
    return target

def _pick_sheet(wb):
    # Prefer a sheet literally named “March 2025”
    for ws in wb.worksheets:
        if _norm(ws.title) == "march 2025":
            return ws
    # Otherwise the active sheet
    return wb.active

def _map_headers(ws):
    """Map logical headers to column indices (1-based) using substrings on row 1."""
    colmap = {}
    for cell in ws[1]:
        h = _norm(cell.value)
        if not h:
            continue
        for key, hints in HEADER_HINTS.items():
            if key in colmap:
                continue
            if any(k in h for k in hints):
                colmap[key] = cell.col_idx
                break
    return colmap

def read_rows(path):
    """
    Return: (wb, ws, rows, colmap)
      rows[i] = {
        "row": int,
        "id_raw": str,        # Column D
        "id_upper": str,
        "description": str,   # from mapped Description header
        "is_p": bool          # Column I ('Yes' => True)
      }
    """
    tmp = _safe_copy_for_read(path)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        try: os.remove(tmp)
        except Exception: pass

    ws = _pick_sheet(wb)
    colmap = _map_headers(ws)

    # Figure out Description col from headers; if not found, fall back near O (15)
    desc_col = colmap.get("description", 15)

    # FORCE Identifier (D) and P applicability (I)
    id_col   = FORCE_IDENTIFIER_COL
    is_p_col = FORCE_ISP_COL

    # Keep colmap in sync (these are where we READ ids/is_p from)
    colmap["identifier"] = id_col
    colmap["is_p"] = is_p_col
    colmap["description"] = desc_col

    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        id_raw  = str(r[id_col - 1].value or "").strip()
        desc    = str(r[desc_col - 1].value or "").strip()
        is_p_v  = r[is_p_col - 1].value if is_p_col <= len(r) else ""
        is_p    = str(is_p_v or "").strip().lower() in ("y", "yes", "true", "1")
        rows.append({
            "row": i,
            "id_raw": id_raw,
            "id_upper": id_raw.upper(),
            "description": desc,
            "is_p": is_p
        })

    return wb, ws, rows, colmap

def write_outcome_pqr(wb, ws, row_index: int, p_entity: str, q_status: str, r_comment: str, colmap: dict):
    """Write only the mapped P/Q/R columns (no debug mirrors)."""
    ws.cell(row=row_index, column=colmap["resp_entity"],   value=(p_entity or ""))
    ws.cell(row=row_index, column=colmap["impl_status"],   value=(q_status or ""))
    ws.cell(row=row_index, column=colmap["impl_comments"], value=(r_comment or ""))
